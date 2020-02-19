import glob
import os
import datetime
import smtplib
import pandas as pd
import openpyxl as xl
import xlsxwriter
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from tkinter import *
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Combine, concatenate, join multiple excel files with multiple sheets in a given folder into one dataframe,
# All sheets in a single Excel file are first combined into a dataframe, then all the Excel Books in the folder
# Are combined to make a single data frame. The combined data frame is the exported into a single Excel sheet.

# path = current working directory
root = Tk()
root.withdraw()

# set file Directory

path = filedialog.askdirectory()

# Set variable today for PO

today = datetime.datetime.today().strftime('%Y%m%d')

masterFile = "%s\\MF_%s.xlsx" % (path, today)
dcMaster = "%s\\DC_StoreMF_%s.xlsx" % (path, today)
amyMaster = "%s\\amyMF_%s.xlsx" % (path, today)


# Combine all current .xlsx files in current working Directory

def concatenate_master():
    try:

        # Check if .xlsx files exist in folder

        xlsxCheck = glob.glob('%s/*.xlsx' % path)

        if len(xlsxCheck) == 0:
            root.withdraw()
            tk.messagebox.showinfo('Error',
                                   'No excel files exist in folder',
                                   icon='warning')
            exit()

        # Grab all files with extension .xlsx

        filenames = glob.glob('%s/*.xlsx' % path)

        # Dataframe Initialization

        concat_all_sheets_all_files = pd.DataFrame()

        for file in filenames:

            # Get all the sheets in a single Excel File using  pd.read_excel command, with sheet_name=None
            # Note that the result is given as an Ordered Dictionary File

            # If MF exists in name, skip file

            if 'MF_' in file:
                continue

            # Filecheck if exists, remove

            elif 'MF_%s' % today in file:
                os.remove(file)

            df = pd.read_excel(
                file,
                sheet_name=None,
                skiprows=None,
                nrows=None,
                usecols=None,
                header=0,
                index_col=None,
            )

            # Use pd.concat command to Concatenate pandas objects as a Single Table.

            concat_all_sheets_single_file = pd.concat(df, sort=False)

            # Use append command to append/stack the previous concatenated data on top of each other
            # as the iteration goes on for every files in the folder

            concat_all_sheets_all_files = concat_all_sheets_all_files.append(concat_all_sheets_single_file)

        # write data to MF.xlsx

        writer = pd.ExcelWriter(masterFile)

        concat_all_sheets_all_files.to_excel(writer)

        writer.save()

    except PermissionError:
        root.withdraw()
        tk.messagebox.showerror('Error',
                                'Please close files in folder before running',
                                icon='error')
        exit()
    unmerge()


# -----

def unmerge():
    # open file MF.xlsx

    wb1 = xl.load_workbook(masterFile)
    ws1 = wb1['Sheet1']

    # set A1 = vendor, B1 = sheetIndex

    ws1['A1'] = 'vendor'
    ws1['B1'] = 'sheetIndex'

    # get ranges of merged cells

    mergedRanges = ws1.merged_cells.ranges

    # unmerge cells

    while mergedRanges:
        for entry in mergedRanges:
            ws1.unmerge_cells(str(entry))

    wb1.save(masterFile)

    # Fill cells with last populated value (pandas)

    fill(masterFile)


# -----

def subMF():
    # Read masterFile

    masterFileData = pd.read_excel(masterFile)

    # Create dc Data Frame

    dcDataFrame = masterFileData.loc[:, [
                                            'vendor',
                                            'UPC',
                                            'Expiration_Date',
                                            'Description',
                                            'Pcs_per_case',
                                            'Pcs_per_Master',
                                            'Order',
                                            'Total_pcs',
                                            'Price',
                                            'Region',
                                            'Store',
                                        ]]

    # Filter out any total_pcs = 0

    dcDataFrame = dcDataFrame[dcDataFrame.Total_pcs != 0]

    # Create WB's (dcMaster, amyMaster)

    createDCwb = xlsxwriter.Workbook(dcMaster)
    createAMYwb = xlsxwriter.Workbook(amyMaster)
    createAMYwb.close()
    createDCwb.close()

    # Write dcDataFrame to excel (dcMaster)

    dcDataFrame.to_excel(dcMaster)

    fill(dcMaster)

    masterFileData['Expiration_Date'] = masterFileData['Expiration_Date'].fillna('')

    # Create amy Data Frame

    amyDF = masterFileData.groupby(['vendor', 'UPC', 'Expiration_Date', 'UnitCost']).agg({'Order': 'sum',
                                                                                          'Total_pcs': 'sum'})

    # Write amy DataFrame to excel (amyMaster)

    pd.DataFrame(amyDF).to_excel(amyMaster)

    # Read amyMaster as new Data Frame

    amyData = pd.read_excel(amyMaster)

    # Calculate Total from unit cost * total pieces

    amyData['Total'] = amyData['UnitCost'] * amyData['Total_pcs']

    # Write dataframe to excel (amyMaster)

    pd.DataFrame(amyData).to_excel(amyMaster)

    fill(amyMaster)


# -----

def fill(filename):
    # Fill in blanks in 'vendor' column w/ last populated cell in column

    df = pd.read_excel(filename)
    df['vendor'].fillna(method='ffill', inplace=True)
    df.to_excel(filename)


# -----

def userprompt():
    # Prompt user to send email or not
    # Process done running, either selection will generate files

    root.withdraw()
    MsgBox = tk.messagebox.askquestion('Send Mail',
                                       'Would you like to send an email?', icon='warning')

    if MsgBox == 'yes':

        sendMail()

        tk.messagebox.showinfo('Success',
                               'Email was sent. Files:\nMF_%s.xlsx\nDC_StoreMF_%s.xlsx, &'
                               '\namyMF_%s.xlsxSuccessfully Generated' % (today, today, today))
        exit()
    else:
        tk.messagebox.showinfo('Success',
                               'No Email was sent. Files:\nMF_%s.xlsx\nDC_StoreMF_%s.xlsx, &'
                               '\namyMF_%s.xlsxSuccessfully Generated' % (today, today, today))
        exit()


# Send email to appropriate users with zip of masterFile

def sendMail():
    fromaddr = 'add email here'
    toaddr = 'add email here'
    files = ['\\MF_%s.xlsx' % today, '\\DC_StoreMF_%s.xlsx' % today, '\\amyMF_%s.xlsx' % today]
    msg = MIMEMultipart()

    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = 'MasterFiles Generated: %s' % today

    body = 'MF_%s.xlsx,\nDC_StoreMF_%s.xlsx, &\namyMF_%s.xlsx Successfully Generated' % (today, today, today)

    msg.attach(MIMEText(body, 'plain'))

    for f in files:  # add files to the message
        attachment = MIMEApplication(open(path + f, 'rb').read(), _subtype='txt')
        attachment.add_header('Content-Disposition', 'attachment', filename=f)
        msg.attach(attachment)

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(fromaddr, 'add password here') # change password to appropriate pw
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr, text)
    server.quit()


concatenate_master()
subMF()
userprompt()
